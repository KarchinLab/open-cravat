import argparse
import os
import subprocess
from cravat import admin_util as au
import time
import sys
from abc import ABC, abstractmethod
import csv as csv
import openpyxl as pyxl
import openpyxl as pyxl


# Regression test program for CRAVAT modules.  By default, it will go through all modules and
# if the module has a test directory with an input and key file, it will test the module
# and compare the results to the key to determine whether everything is running as expected.
# The command line -m can be used to specify one or more specific modules to test and the
# -t option can be used to run tests on modules of a specific type (e.g. annotator)
#
# The test program runs input though the full cravat processing flow and collects the output
# of the 'text' reporter to compare to the key.
#
# When creating a test case for the cravat_test, just run a cravat input, collect the text
# report output, check that the results are correct, and then save the output as the key.
#
# The tester creates an output directory and then a subdirectory for each module run.
# Logs and output for each module can be found in the associated subdirectory.


# Open CRAVAT has multiple output types (reports) generated optional using the -t parameter when running oc.
# The tester uses report output to verify expected results.  The ReportReader class is responsible for parsing
# output reports and 'key' files which are expected results.  Because there are multiple types of reports,
# the ReportReader is an abstract class and a class of the correct type is instantiated based on the type of
# report selected in the test.
class ReportReader(ABC):
    def __init__(self, rsltFile):
        self.rsltFile = rsltFile

    # Return the file extension of this type of report (e.g. text is .tsv)
    @abstractmethod
    def reportFileExtension(self):
        pass

    # Read the specified level (variant, gene, etc) from a cravat report
    # Return a list of the report headers and a list of the row values
    # The bDict parameter indicates whether to return Rows as a list or dictionary
    @abstractmethod
    def readReport(self, test_level, bDict):
        pass


# Derived Report Reader class for reating text reports (-t text)
class TextReportReader(ReportReader):
    def reportFileExtension(self):
        return ".tsv"

    # Based on the level selected, return column headers and row values.
    def readReport(self, test_level, bDict):
        level_hdr = "Report level:"
        level = ""
        headers = None
        if bDict:
            rows = {}
        else:
            rows = []
        with open(self.rsltFile, encoding="latin-1") as f:
            line = f.readline().strip("\n")
            while line:
                # skip comment lines but pull out the report level
                if line.strip().startswith("#"):
                    if level_hdr in line:
                        level = line[line.index(level_hdr) + len(level_hdr) + 1 :]
                    line = f.readline().strip("\n")
                    continue

                # only load the level we are testing
                if level != test_level:
                    line = f.readline().strip("\n")
                    continue

                # load headers for the section
                if headers == None:
                    line2 = f.readline().strip("\n")
                    headers = self.readSectionHeader(line, line2)
                    line = f.readline().strip("\n")
                    continue

                columns = line.split("\t")
                line_id = self.getRowID(headers, columns, level)
                if bDict:
                    rows[line_id] = columns
                else:
                    rows.append((line_id, columns))
                line = f.readline().strip("\n")
        return headers, rows

    # Read the two report header columns that define the module/column
    # for each data column.  Returned as list of: module|column
    def readSectionHeader(self, line1, line2):
        cols = line1.split("\t")
        headers = []
        current_module = cols[0]
        for col in cols:
            if col != "":
                current_module = col
                headers.append(col + "|")
            else:
                headers.append(current_module + "|")
        cols = line2.split("\t")
        for idx, col in enumerate(cols):
            headers[idx] = headers[idx] + col
        return headers

    # The ID of a result row is used to match key and output.  The ID
    # differs depending on which section of the output is being checked.
    def getRowID(self, headers, columns, level):
        Id = ""
        if level == "variant":
            Id = (
                columns[self.getColPos(headers, "Chrom")]
                + " "
                + columns[self.getColPos(headers, "Position")]
                + " "
                + columns[self.getColPos(headers, "Ref Base")]
                + " "
                + columns[self.getColPos(headers, "Alt Base")]
                + " "
                + columns[self.getColPos(headers, "Tags")]
            )
        if level == "gene":
            pos = self.getColPos(headers, "Hugo")
            if pos == -1:
                pos = self.getColPos(headers, "Gene")
            Id = columns[pos]
        if level == "sample":
            Id = (
                columns[self.getColPos(headers, "UID")]
                + " "
                + columns[self.getColPos(headers, "Sample")]
            )
        if level == "mapping":
            Id = columns[self.getColPos(headers, "Original Line")]
        return Id

    # get the position of a specific output column
    def getColPos(self, headers, col):
        for idx, header in enumerate(headers):
            if col in header:
                return idx
        return -1


# Derived Report Reader class for reating text reports (-t text)
class ExcelReportReader(ReportReader):
    def reportFileExtension(self):
        return ".xlsx"

    # Based on the level selected, return column headers and row values.
    def readReport(self, test_level, bDict):
        headers = None
        tabNbr = "Variant"
        if test_level == "gene":
            tabNbr = "Gene"
        elif test_level == "sample":
            tabNbr = "Sample"
        elif test_level == "mapping":
            tabNbr = "Mapping"

        # To open Workbook
        xlxsFile = (
            self.rsltFile if ".xlsx" in self.rsltFile else self.rsltFile + ".xlsx"
        )
        wb = pyxl.load_workbook(filename=xlxsFile)
        sheet = wb[tabNbr]

        if bDict:
            rows = {}
        else:
            rows = []
        if headers == None:
            headers = self.readSectionHeader(test_level, sheet)
        for i in range(3, sheet.max_row + 1):
            columns = []
            for j in range(1, sheet.max_column + 1):
                columns.append(
                    "" if sheet.cell(i, j).value is None else sheet.cell(i, j).value
                )
            line_id = self.getRowID(headers, columns, test_level)
            if bDict:
                rows[line_id] = columns
            else:
                rows.append((line_id, columns))
        return headers, rows

    # Read the two report header columns that define the module/column
    # for each data column.  Returned as list of: module|column
    def readSectionHeader(self, test_level, sheet):
        headers = []
        # To open Workbook
        header1 = sheet.cell(1, 1).value
        for i in range(1, sheet.max_column + 1):
            if sheet.cell(1, i).value is not None and sheet.cell(1, i).value != "":
                header1 = sheet.cell(1, i).value
            header2 = sheet.cell(2, i).value
            combinedHeader = header1 + "|" + header2
            headers.append(combinedHeader)
        return headers

    # The ID of a result row is used to match key and output.  The ID
    # differs depending on which section of the output is being checked.
    def getRowID(self, headers, columns, level):
        Id = ""
        if level == "variant":
            Id = (
                columns[self.getColPos(headers, "Chrom")]
                + " "
                + str(int(columns[self.getColPos(headers, "Position")]))
                + " "
                + columns[self.getColPos(headers, "Ref Base")]
                + " "
                + columns[self.getColPos(headers, "Alt Base")]
                + " "
                + columns[self.getColPos(headers, "Tags")]
            )
        if level == "gene":
            pos = self.getColPos(headers, "Hugo")
            if pos == -1:
                pos = self.getColPos(headers, "Gene")
            Id = columns[pos]
        if level == "sample":
            Id = (
                columns[self.getColPos(headers, "UID")]
                + " "
                + columns[self.getColPos(headers, "Sample")]
            )
        if level == "mapping":
            Id = columns[self.getColPos(headers, "Original Line")]
        return Id

    # get the position of a specific output column
    def getColPos(self, headers, col):
        for idx, header in enumerate(headers):
            if col in header:
                return idx
        return -1


# Derived Report Reader class for reating text reports (-t text)
class VcfReportReader(ReportReader):
    def reportFileExtension(self):
        return ".vcf"

    # Based on the level selected, return column headers and row values.
    def readReport(self, test_level, bDict):
        headers = None
        if bDict:
            rows = {}
        else:
            rows = []
        import vcf
        reader = vcf.Reader(filename=self.rsltFile)
        if headers == None:
            headers = self.readSectionHeader(reader)
        for record in reader:
            columns = []
            columns.append(record.ID)
            lineitems = record.INFO["CRV"][0].split("|")
            i = 1
            while i < len(headers):
                columns.append(lineitems[i].strip('"'))
                i += 1
                line_id = self.getRowID(headers, lineitems)
            if bDict:
                rows[line_id] = columns
            else:
                rows.append((line_id, columns))
        return headers, rows

    # Read the two report header columns that define the module/column
    # for each data column.  Returned as list of: module|column
    def readSectionHeader(self, reader):
        headers = [
            "Variant Annotation|chrom",
            "Variant Annotation|pos",
            "Variant Annotation|ref",
            "Variant Annotation|alt",
        ]
        colList = reader.infos["CRV"][3]
        cols = colList.split("|")
        for col in cols:
            underscorepos = col.find("__")
            equalspos = col.find("=")
            prevheader = col[underscorepos - 4 : underscorepos]
            postheader = col[underscorepos + 2 : len(col)]
            hugopos = postheader.find("hugo")
            if hugopos >= 0:
                postheader = "hugo"
            elif equalspos > 0:
                break
            if prevheader == "base":
                header = "Variant Annotation|" + postheader
            else:
                module = col[0:underscorepos]
                header = module + "|" + postheader
            headers.append(header)
        return headers

    # The ID of a result row is used to match key and output.  The ID
    # differs depending on which section of the output is being checked.
    def getRowID(self, headers, lineitems):
        Id = (
            lineitems[self.getColPos(headers, "Variant Annotation|chrom")].strip('"')
            + " "
            + lineitems[self.getColPos(headers, "Variant Annotation|pos")].strip('"')
            + " "
            + lineitems[self.getColPos(headers, "Variant Annotation|ref")].strip('"')
            + " "
            + lineitems[self.getColPos(headers, "Variant Annotation|alt")].strip('"')
            + " "
            + lineitems[self.getColPos(headers, "Variant Annotation|tags")].strip('"')
        )
        return Id

    # get the position of a specific output column
    def getColPos(self, headers, col):
        for idx, header in enumerate(headers):
            if col in header:
                return idx
        return -1


# Derived Report Reader class for reating text reports (-t text)
class TsvReportReader(ReportReader):
    def reportFileExtension(self):
        return ".variant.tsv"

    # Based on the level selected, return column headers and row values.
    def readReport(self, test_level, bDict):
        level_hdr = "level="
        level = ""
        headers = None
        if bDict:
            rows = {}
        else:
            rows = []
        with open(self.rsltFile, encoding="latin-1") as f:
            line = f.readline().strip("\n")
            while line:
                # skip comment lines but pull out the report level
                if line.strip().startswith("#"):
                    if level_hdr in line:
                        level = line[line.index(level_hdr) + len(level_hdr) :]
                    line = f.readline().strip("\n")
                    continue

                # only load the level we are testing
                if level != test_level:
                    line = f.readline().strip("\n")
                    continue

                # load headers for the section
                if headers == None:
                    headers = self.readSectionHeader(line)
                    line = f.readline().strip("\n")
                    continue

                columns = line.split("\t")
                line_id = self.getRowID(headers, columns, level)
                if bDict:
                    rows[line_id] = columns
                else:
                    rows.append((line_id, columns))
                line = f.readline().strip("\n")
        return headers, rows

    # Read the two report header columns that define the module/column
    # for each data column.  Returned as list of: module|column
    def readSectionHeader(self, line):
        cols = line.split("\t")
        headers = []
        for col in cols:
            header = col.replace(".", "|", 1)
            if not "|" in header:
                header = "Variant Annotation|" + header
            headers.append(header)
        return headers

    # The ID of a result row is used to match key and output.  The ID
    # differs depending on which section of the output is being checked.
    def getRowID(self, headers, columns, level):
        Id = ""
        if level == "variant":
            Id = (
                columns[self.getColPos(headers, "chrom")]
                + " "
                + columns[self.getColPos(headers, "pos")]
                + " "
                + columns[self.getColPos(headers, "ref_base")]
                + " "
                + columns[self.getColPos(headers, "alt_base")]
                + " "
                + columns[self.getColPos(headers, "tags")]
            )
        if level == "gene":
            pos = self.getColPos(headers, "Hugo")
            if pos == -1:
                pos = self.getColPos(headers, "Gene")
            Id = columns[pos]
        if level == "sample":
            Id = (
                columns[self.getColPos(headers, "UID")]
                + " "
                + columns[self.getColPos(headers, "Sample")]
            )
        if level == "mapping":
            Id = columns[self.getColPos(headers, "Original Line")]
        return Id

    # get the position of a specific output column
    def getColPos(self, headers, col):
        for idx, header in enumerate(headers):
            if col in header:
                return idx
        return -1


# Derived Report Reader class for reating text reports (-t text)
class CsvReportReader(ReportReader):
    def reportFileExtension(self):
        return ".variant.csv"

    # Based on the level selected, return column headers and row values.
    def readReport(self, test_level, bDict):
        level_hdr = "level="
        level = ""
        headers = None
        if bDict:
            rows = {}
        else:
            rows = []
        with open(self.rsltFile, encoding="latin-1") as f:
            rdr = csv.reader(f)
            for row in rdr:
                # skip comment lines but pull out the report level
                if row[0].startswith("#"):
                    if level_hdr in row[0]:
                        hdr_line = row[0]
                        level = hdr_line[hdr_line.index(level_hdr) + len(level_hdr) :]
                    continue

                # only load the level we are testing
                if level != test_level:
                    continue

                # load headers for the section
                if headers == None:
                    headers = self.readSectionHeader(row)
                    continue

                # load headers for the section
                line_id = self.getRowID(headers, row, level)
                if bDict:
                    rows[line_id] = row
                else:
                    rows.append((line_id, row))
        return headers, rows

    # Read the two report header column that define the module/column
    # for each data column.  Returned as list of: module|column
    def readSectionHeader(self, row):
        headers = []
        for col in row:
            # tester use '|' and csv uses ':', just switch the character
            header = col.replace(".", "|", 1)
            # tester is expecting base module to be 'Variant Annotation' so switch it.
            if not "|" in header:
                header = "Variant Annotation|" + header

            headers.append(header)
        return headers

    # The ID of a result row is used to match key and output.  The ID
    # differs depending on which section of the output is being checked.
    def getRowID(self, headers, columns, level):
        Id = ""
        if level == "variant":
            Id = (
                columns[self.getColPos(headers, "chrom")]
                + " "
                + columns[self.getColPos(headers, "pos")]
                + " "
                + columns[self.getColPos(headers, "ref_base")]
                + " "
                + columns[self.getColPos(headers, "alt_base")]
                + " "
                + columns[self.getColPos(headers, "tags")]
            )
        if level == "gene":
            pos = self.getColPos(headers, "Hugo")
            if pos == -1:
                pos = self.getColPos(headers, "Gene")
            Id = columns[pos]
        if level == "sample":
            Id = (
                columns[self.getColPos(headers, "UID")]
                + " "
                + columns[self.getColPos(headers, "Sample")]
            )
        if level == "mapping":
            Id = columns[self.getColPos(headers, "Original Line")]
        return Id

    # get the position of a specific output column
    def getColPos(self, headers, col):
        for idx, header in enumerate(headers):
            if col in header:
                return idx
        return -1


# class that actually runs a test of a specific module and then verifies the results.
class Tester:
    def __init__(self, module, rundir, input_file):
        cur_dir = os.path.dirname(os.path.abspath(__file__))
        self.module = module
        if not os.path.exists(module.directory) or not module.script_exists:
            raise Exception(
                "No runnable module installed at path %s" % module.directory
            )
        self.out_dir = os.path.join(rundir, module.name)
        if not os.path.exists(self.out_dir):
            os.makedirs(self.out_dir)
        self.input_file = input_file
        self.input_path = os.path.join(module.test_dir, input_file)
        self.key_path = os.path.join(
            module.test_dir, input_file.replace("input", "key")
        )
        self.parms_path = os.path.join(
            module.test_dir, input_file.replace("input", "parms")
        )
        log = "test.log"
        if len(input_file.replace("input", "")) > 0:
            log = input_file + ".test.log"
        self.log_path = os.path.join(
            self.out_dir, log
        )  # put the output of this program in test.log
        self.cravat_run = os.path.join(cur_dir, "runcravat.py")
        self.output_file = "oc_output"
        self.out_path = os.path.join(self.out_dir, self.output_file)
        self.log = open(self.log_path, "w", encoding="UTF-8")
        self.start_time = None
        self.end_time = None
        self.failures = []
        self.test_passed = False
        self.report_type = "text"

    # optionally a test director for a module can have a 'parms' file.  If it does,
    # the test parameters are tab delimited.  Load each parm/value into the parms
    # dictionary.
    def parse_parms(self):
        self.parms = {}
        if os.path.exists(self.parms_path):
            with open(self.parms_path) as f:
                line = f.readline().strip("\n")
                while line:
                    parm = line.split("\t")
                    if len(parm) == 2:
                        # put the parameter and value in the parms dictionary
                        self.parms[parm[0]] = parm[1]
                    line = f.readline().strip("\n")

    # function that tests one module
    def run(self):
        input_msg = (
            "" if self.input_file == "input" else self.input_file
        )  # if there is more than one test for the module, include the test file in the log.
        self._report("  Testing: " + self.module.name + " " + input_msg)
        self.start_time = time.time()
        self.parse_parms()
        python_exc = sys.executable

        # default is to run 'text' report but it can be overridden in the optional parms file.
        if "Report_Type" in self.parms:
            self.report_type = self.parms["Report_Type"]
        else:
            self.report_type = "text"

        # Basic oc run command line
        cmd_list = [
            python_exc,
            self.cravat_run,
            self.input_path,
            "-d",
            self.out_dir,
            "-n",
            self.output_file,
            "-t",
            self.report_type,
        ]
        if self.module.type == "annotator":
            cmd_list.extend(["-a", self.module.name])
        elif (
            (self.module.type == "reporter")
            and (au.get_local_module_info("vest") is not None)
            and (au.get_local_module_info("cgl") is not None)
        ):
            # when testing reporters, if the vest and cgl modules are installed, include them in the run / report.
            cmd_list.extend(["-a", "vest", "cgl"])
        else:
            cmd_list.extend(["--skip", "annotator"])

        # special case for a few converter modules that need hg19 coordinates
        if self.module.name in [
            "ftdna-converter",
            "ancestrydna-converter",
            "23andme-converter",
        ]:
            cmd_list.extend(["-l", "hg19"])
        else:
            cmd_list.extend(["-l", "hg38"])

        print(" ".join(cmd_list))
        exit_code = subprocess.call(
            " ".join(cmd_list), shell=True, stdout=self.log, stderr=subprocess.STDOUT
        )
        if exit_code != 0:
            self._report("    CRAVAT non-zero exit code: " + str(exit_code))
        return exit_code

    def verify(self):
        self.test_passed = True
        if self.module.type == "annotator":
            self.verify_level(self.module.level, [self.module.title])
        elif self.module.type == "converter":
            self.verify_level("variant", ["Variant Annotation"])
            self.verify_level("sample", ["Variant Annotation"])
            self.verify_level("mapping", ["Variant Annotation"])
        elif self.module.type == "mapper":
            self.verify_level("variant", ["Variant Annotation"])
            self.verify_level("gene", ["Variant Annotation"])
        elif self.module.type == "reporter":
            if self.report_type == "vcf":
                self.verify_level(
                    "variant",
                    [
                        "Variant Annotation",
                        "vest",
                        "cgl",
                        "VEST4",
                        "Cancer Gene Landscape",
                    ],
                )
            else:
                if (au.get_local_module_info("vest") is not None) and (
                    au.get_local_module_info("cgl") is not None
                ):
                    self.verify_level(
                        "variant",
                        [
                            "Variant Annotation",
                            "vest",
                            "cgl",
                            "VEST4",
                            "Cancer Gene Landscape",
                        ],
                    )
                    self.verify_level(
                        "gene",
                        [
                            "Variant Annotation",
                            "vest",
                            "cgl",
                            "VEST4",
                            "Cancer Gene Landscape",
                        ],
                    )
                else:
                    self.verify_level("variant", ["Variant Annotation"])
                    self.verify_level("gene", ["Variant Annotation"])

    # See if key and result are floating point numbers.  If so, allow tiny
    # differences.
    def floats_differ(self, str_val1, str_val2):
        try:
            v1 = float(str_val1)
            v2 = float(str_val2)
        except:
            return True

        if abs(v1 - v2) < 0.002:
            return False
        else:
            return True

    # based on the type of report run in this test, create the appropriate type of
    # report reader.
    def create_report_reader(self, type, report_path):
        if type == "text":
            return TextReportReader(report_path)
        elif type == "tsv":
            return TsvReportReader(report_path)
        elif type == "csv":
            return CsvReportReader(report_path)
        elif type == "excel":
            return ExcelReportReader(report_path)
        elif type == "vcf":
            return VcfReportReader(report_path)

        # need to put more parsers here when they are implemented

    # Match the key (expected values) to the text report output.  Generate errors
    # if expected results are not found and fail the test.    Test just the specified
    # level (variant, gene, etc) and specified module's columns
    def verify_level(self, level, module_name):
        self._report("  Verifying " + level + " level values.")
        key_reader = self.create_report_reader(self.report_type, self.key_path)
        report_extension = key_reader.reportFileExtension()
        result_reader = self.create_report_reader(
            self.report_type, self.out_path + report_extension
        )
        key_header, key_rows = key_reader.readReport(level, False)
        result_header, result_rows = result_reader.readReport(level, True)
        for key in key_rows:
            variant, key_row = key
            if variant not in result_rows:
                self._report("  Variant: " + variant + " did not appear in results")
                self.test_passed = False
                continue

            result = result_rows[variant]

            for idx, header in enumerate(key_header):
                # just check the columns from the module we are testing
                if (
                    (self.getModule(header) not in module_name)
                    or "uid" in header
                    or "UID" in header
                ):
                    continue

                if header not in result_header:
                    self._report(
                        "  Expected Header: " + header + " did not appear in results"
                    )
                    self.test_passed = False
                    continue

                result_idx = result_header.index(header)
                if (result[result_idx] != key_row[idx]) and self.floats_differ(
                    result[result_idx], key_row[idx]
                ):
                    headLabel = header
                    if "|" in header:
                        headLabel = header[header.index("|") + 1 :]
                    self._report(
                        "      "
                        + variant
                        + "  "
                        + headLabel
                        + "  Expected: "
                        + key_row[idx]
                        + "  Result: "
                        + result[result_idx]
                    )
                    self.test_passed = False

    # headers are <module name>|<header> - this extracts the module name
    def getModule(self, header):
        return header[: header.index("|")]

    # Write a message to the screen and to the log file.
    def _report(self, s):
        self.log.write(s + "\n")
        print(s)

    # Log success /failure of test.
    def write_results(self):
        self.end_time = time.time()
        elapsed_time = self.end_time - self.start_time
        self._report("  Completed in: %.2f seconds" % elapsed_time)
        if self.test_passed:
            self._report("  Test result: PASS")
        else:
            self._report("  Test result: FAIL")


parser = argparse.ArgumentParser()
parser.add_argument("-d", "--rundir", help="Directory for output")
parser.add_argument(
    "-m", "--modules", nargs="+", help="Name of module(s) to test. (e.g. gnomad)"
)
parser.add_argument(
    "-t", "--mod_types", nargs="+", help="Type of module(s) to test (e.g. annotators)"
)

# Loop through installed modules.  Test each one or the ones indicated
# by -m and -t options.
def run_test(cmd_args):
    if cmd_args.rundir is None:
        cmd_args.rundir = "cravat_test_" + str(int(round(time.time() * 1000)))

    # create run output directory
    if not os.path.exists(cmd_args.rundir):
        os.makedirs(cmd_args.rundir)

    # installed module types
    module_types = au.get_local_module_types()

    passed = 0
    failed = 0
    modules_failed = []
    for mod_type in module_types:
        if cmd_args.mod_types is None or mod_type in cmd_args.mod_types:
            print("\nRunning " + mod_type + " tests.")
            modules = au.get_local_module_infos_of_type(mod_type)
            for mod_name in modules:
                if cmd_args.modules is None or mod_name in cmd_args.modules:
                    module = modules[mod_name]
                    # If a module has a test, it is usually a single 'input' file and 'key' but modules can
                    # have multiple input and key files.  This loop runs all input/key file pairs.
                    # Example input.1, key.1, input.2, key.2
                    for test_input_file in module.tests:
                        tester = Tester(module, cmd_args.rundir, test_input_file)
                        exit_code = tester.run()
                        if exit_code == 0:
                            tester.verify()

                        tester.write_results()
                        if tester.test_passed:
                            passed += 1
                        else:
                            failed += 1
                            fail_msg = mod_name + (
                                ""
                                if test_input_file == "input"
                                else " " + test_input_file
                            )
                            modules_failed.append(fail_msg)
    modules_failed.sort()
    print(
        "\nTests complete.  Passed: "
        + str(passed)
        + "  Failed: "
        + str(failed)
        + " ["
        + ", ".join(modules_failed)
        + "]"
    )


def main():
    cmd_args = parser.parse_args()
    run_test(cmd_args)


parser.set_defaults(func=run_test)

if __name__ == "__main__":
    main()
